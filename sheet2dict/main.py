from collections import defaultdict

from openpyxl import load_workbook  # type: ignore
from typing import List, Union

from openpyxl.worksheet.worksheet import Worksheet


class Worksheet:
    def __init__(self) -> None:
        self.sheet_items: Union[List[dict], defaultdict[list]] = []

    def __repr__(self) -> str:
        return str(self.sheet_items)

    def xlsx_to_dict(self, path, select_sheet=None, parse_all_sheets=False):
        """
        Read a Worksheet and return it as array of dictionaries
        :param self: Worksheet Object
        :param path: Path to XLSX file
        :param select_sheet: Set active sheet by name
        :param parse_all_sheets: Use if you want to get data from all of the sheets
        :return: Array of rows as dictionaries or dictionary with arrays as values if parse_all_sheets flag set to True
        """
        book = load_workbook(path)
        self.sheet_items = []
        if parse_all_sheets:
            self.sheet_items = defaultdict(list)

            for sheet_name in book.sheetnames:
                self._parse_sheet_items(book[sheet_name], parsing_all_sheets=True)
            return self

        if select_sheet is None:
            sheet = book.active
        else:
            sheet = book[select_sheet]

        self._parse_sheet_items(sheet)

        return self

    def _parse_sheet_items(self, sheet: Worksheet, parsing_all_sheets: bool = False):
        rows = sheet.max_row
        cols = sheet.max_column
        sheet_title = sheet.title

        def item(row, col):
            return (
                sheet.cell(row=1, column=col).value,
                str(sheet.cell(row=row, column=col).value),
            )

        list_to_append = (
            self.sheet_items[sheet_title] if parsing_all_sheets else self.sheet_items
        )
        for row in range(2, rows + 1):
            list_to_append.append(dict(item(row, col) for col in range(1, cols + 1)))

    def csv_to_dict(self, csv_file, delimiter=","):
        """
        Read a CSV and return it as an array of dictionaries
        :param self: CSV Object
        :param csv_file: CSV file
        :param delimiter: CSV delimiter fe. ';', ','
        :return: Array of rows as dictionaries
        """
        import csv

        self.sheet_items = []

        dict_reader = csv.DictReader(csv_file, delimiter=delimiter)
        self.sheet_items = list(dict_reader)
        return self.sheet_items

    @property
    def header(self):
        """
        Return first row of data as dictionary
        :return: first row of data as dictionary
        """
        return self.sheet_items[0]

    @property
    def sanitize_sheet_items(self):
        """
        Remove None or empty keys from dictionaries
        :param self: XLSX or CSV Object
        :return: Array of rows as dictionaries
        """
        sanitized_items = self.sheet_items
        for item in sanitized_items:
            if None in item:
                del item[None]
            elif "" in item:
                del item[""]
        return sanitized_items
